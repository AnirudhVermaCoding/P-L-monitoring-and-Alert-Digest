"""Build a polished, colour-coded Excel workbook of the computed P&L and anomalies log.

CSV is raw text and can't be styled, so the "beautiful" download is an .xlsx: a dark bold
header row (frozen), sensible column widths, and the Colour column filled with its actual
Red/Yellow/Green/Blue so a non-technical reader sees the day's status at a glance.
"""
from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ARGB fills (no leading '#') matching the app/email palette.
COLOUR_FILL = {"Red": "E74C3C", "Yellow": "F1C40F", "Green": "2ECC71", "Blue": "3498DB"}
# Soft row tints — same hue, easy to read black text on.
COLOUR_TINT = {"Red": "FBE3E3", "Yellow": "FDF3D6", "Green": "E4F6EC", "Blue": "E5EFFB"}
_HEADER_FILL = PatternFill("solid", fgColor="1F2933")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _style_sheet(ws, df: pd.DataFrame, tint_rows: bool = False) -> None:
    """Header styling, frozen top row, column widths, Colour-column fills, optional row tint."""
    for col_idx, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    for col_idx, col in enumerate(df.columns, start=1):
        body = [len(str(v)) for v in df[col].tolist()] if len(df) else [0]
        width = min(max(len(str(col)), max(body)) + 2, 42)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(width, 10)

    if "Colour" not in df.columns:
        return
    ci = list(df.columns).index("Colour") + 1
    n_cols = len(df.columns)
    for row_idx, val in enumerate(df["Colour"].tolist(), start=2):
        if tint_rows and val in COLOUR_TINT:
            tint = PatternFill("solid", fgColor=COLOUR_TINT[val])
            for c in range(1, n_cols + 1):
                ws.cell(row=row_idx, column=c).fill = tint
        if val in COLOUR_FILL:  # the Colour cell itself gets the solid, bold chip
            cell = ws.cell(row=row_idx, column=ci)
            cell.fill = PatternFill("solid", fgColor=COLOUR_FILL[val])
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")


def _empty_like(df: pd.DataFrame) -> pd.DataFrame:
    return df if not df.empty else pd.DataFrame(columns=list(df.columns) or ["Date"])


def build_report_xlsx(pnl_df: pd.DataFrame, anomalies_df: pd.DataFrame) -> bytes:
    """Return a styled two-sheet workbook (P&L + Anomalies) as bytes, ready to download."""
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xl:
        pnl_df.to_excel(xl, sheet_name="P&L", index=False)
        _empty_like(anomalies_df).to_excel(xl, sheet_name="Anomalies", index=False)
        _style_sheet(xl.sheets["P&L"], pnl_df)
        _style_sheet(xl.sheets["Anomalies"], anomalies_df, tint_rows=True)
    return buf.getvalue()


def build_anomalies_xlsx(anomalies_df: pd.DataFrame) -> bytes:
    """Return a single-sheet, colour-coded anomalies workbook — each row tinted by the day's
    colour, with the Colour column shown as a solid Red/Yellow/Green/Blue chip."""
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xl:
        _empty_like(anomalies_df).to_excel(xl, sheet_name="Anomalies", index=False)
        _style_sheet(xl.sheets["Anomalies"], anomalies_df, tint_rows=True)
    return buf.getvalue()
