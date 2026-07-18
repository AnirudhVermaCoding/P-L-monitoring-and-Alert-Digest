"""Tests for the colour-coded Excel report builder. Run: `python -m tests.test_report`."""
from __future__ import annotations

import sys
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from core.config import load_config
from core.engine import compute_pnl, detect_anomalies
from core.report import COLOUR_FILL, build_report_xlsx


def _sample_pnl():
    cfg = load_config()
    rev = 100000
    df = pd.DataFrame([
        {"Date": "2026-07-17", "FC": "FC-A", "Revenue": rev,
         "Manpower": 0.30 * rev, "Packaging": 0.08 * rev, "Power and Fuel": 0.16 * rev,
         "FC Rent": 0.18 * rev, "Equipment Rentals": 0.07 * rev, "Overheads": 0.13 * rev},
        {"Date": "2026-07-17", "FC": "FC-B", "Revenue": rev,
         "Manpower": 0.26 * rev, "Packaging": 0.06 * rev, "Power and Fuel": 0.11 * rev,
         "FC Rent": 0.165 * rev, "Equipment Rentals": 0.065 * rev, "Overheads": 0.11 * rev},
    ])
    return compute_pnl(df, cfg), cfg


def test_report_has_two_styled_sheets():
    pnl, cfg = _sample_pnl()
    anoms = detect_anomalies(pnl, cfg)
    data = build_report_xlsx(pnl, anoms)
    assert isinstance(data, bytes) and len(data) > 0

    wb = load_workbook(BytesIO(data))
    assert wb.sheetnames == ["P&L", "Anomalies"], wb.sheetnames

    # Header row is bold on both sheets, and the top row is frozen.
    for name in wb.sheetnames:
        ws = wb[name]
        assert ws["A1"].font.bold, f"{name} header not bold"
        assert ws.freeze_panes == "A2", f"{name} not frozen"

    # The Anomalies sheet has a Colour column whose cells are filled with the colour.
    ws = wb["Anomalies"]
    headers = [c.value for c in ws[1]]
    assert "Colour" in headers, headers
    ci = headers.index("Colour") + 1
    filled = 0
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=ci).value
        if val in COLOUR_FILL:
            rgb = (ws.cell(row=row, column=ci).fill.fgColor.rgb or "")
            assert rgb.endswith(COLOUR_FILL[val]), (val, rgb)
            filled += 1
    assert filled > 0, "no colour fills applied"
    print(f"PASS test_report_has_two_styled_sheets ({filled} colour cells)")


def test_report_handles_empty_anomalies():
    pnl, cfg = _sample_pnl()
    empty = detect_anomalies(pnl.iloc[0:0].assign(Colour=[]), cfg) if False else \
        detect_anomalies(pnl, cfg).iloc[0:0]
    data = build_report_xlsx(pnl, empty)  # must not raise
    wb = load_workbook(BytesIO(data))
    assert wb.sheetnames == ["P&L", "Anomalies"]
    print("PASS test_report_handles_empty_anomalies")


def run_all():
    test_report_has_two_styled_sheets()
    test_report_handles_empty_anomalies()
    print("\nALL REPORT TESTS PASSED")


if __name__ == "__main__":
    run_all()
